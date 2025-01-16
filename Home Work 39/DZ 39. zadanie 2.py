from openpyxl import load_workbook
import numpy as np

def data(file):
    sheet = file.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

def sales(data):
    sr_znach = np.mean(data)
    sr_vzvesh = np.average(data)
    max_znach = np.max(data)
    min_znach = np.min(data)
    summa = np.sum(data)

    print("Среднее значение:", sr_znach)
    print("Средневзвешенное значение:", sr_vzvesh)
    print("Максимальное значение:", max_znach)
    print("Минимальное значение:", min_znach)
    print("Сумма всех значений:", summa)


file = load_workbook('валюты.xlsx')
numpy = np.array(data(file))
sales(numpy)

