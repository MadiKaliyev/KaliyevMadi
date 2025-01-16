from openpyxl import load_workbook
import numpy as np

def data(file):
    sheet = file.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

file = load_workbook('валюты.xlsx')
numpy = np.array(data(file))
print(numpy)
print(*numpy)

