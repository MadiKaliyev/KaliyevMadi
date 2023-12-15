# а) Создайте excel файл в операционной системе, заполните его данными в одну строку,
# от 1 до 10. Во второй строке от а и до 10 буквы.
# б) Прочитайте эти две строки.
# в) Перезапишите их в другой файл, но вертикально, т.е. первая строка – первый столбец
import sys
import os
import string
from openpyxl import Workbook, load_workbook
import openpyxl

wb = Workbook()
ws = wb.create_sheet('книга',0)
fn = 'файл.xlsx'

date = list(range(1, 11))
letters = list(string.ascii_uppercase)[:10]
ws.append(letters)
ws.append(date)

wb.save(filename=fn)
wb.close()

wb1 = Workbook()
ws1 = wb1.create_sheet('книга1',0)
fn_new = 'новый файл.xlsx'

for i in range(1,11):
    ws1.cell(row=i, column=1, value=i)

for i in range(1,11):
    ws1.cell(row=i, column=2, value=chr(64 + i))

wb1.save(filename=fn_new)
wb1.close()



