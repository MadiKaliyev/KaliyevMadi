from tkinter import *
from tkinter import messagebox, simpledialog, filedialog
from PIL import Image, ImageTk
import mysql.connector
import base64
import io
from openpyxl import load_workbook

class Contact:
    def __init__(self, id, l_name, f_name, phone_number, m_name='', birthday='', email='', comment='', relationship='', photo=None):
        self.id = id
        self.l_name = l_name
        self.f_name = f_name
        self.m_name = m_name
        self.phone_number = phone_number
        self.birthday = birthday
        self.email = email
        self.comment = comment
        self.relationship = relationship
        self.photo = photo

    def toStr(self):
        return f"{self.f_name} {self.l_name[0]}. {self.phone_number} {self.comment}"

class EditContactDialog(simpledialog.Dialog):
    def __init__(self, parent, contact):
        self.contact = contact
        super().__init__(parent)

    def body(self, master):
        Label(master, text="Фамилия:").grid(row=0, sticky=W)
        Label(master, text="Имя:").grid(row=1, sticky=W)
        Label(master, text="Номер телефона:").grid(row=2, sticky=W)
        Label(master, text="Дата рождения:").grid(row=3, sticky=W)
        Label(master, text="Email:").grid(row=4, sticky=W)
        Label(master, text="Родственность:").grid(row=5, sticky=W)

        self.l_name_entry = Entry(master)
        self.f_name_entry = Entry(master)
        self.phone_entry = Entry(master)
        self.birthday_entry = Entry(master)
        self.email_entry = Entry(master)
        self.relationship_entry = Entry(master)

        self.l_name_entry.grid(row=0, column=1)
        self.f_name_entry.grid(row=1, column=1)
        self.phone_entry.grid(row=2, column=1)
        self.birthday_entry.grid(row=3, column=1)
        self.email_entry.grid(row=4, column=1)
        self.relationship_entry.grid(row=5, column=1)

        self.l_name_entry.insert(0, self.contact.l_name)
        self.f_name_entry.insert(0, self.contact.f_name)
        self.phone_entry.insert(0, self.contact.phone_number)
        self.birthday_entry.insert(0, self.contact.birthday)
        self.email_entry.insert(0, self.contact.email)
        self.relationship_entry.insert(0, self.contact.relationship)

    def apply(self):
        self.contact.l_name = self.l_name_entry.get()
        self.contact.f_name = self.f_name_entry.get()
        self.contact.phone_number = self.phone_entry.get()
        self.contact.birthday = self.birthday_entry.get()
        self.contact.email = self.email_entry.get()
        self.contact.relationship = self.relationship_entry.get()

class AddContactDialog(simpledialog.Dialog):
    def __init__(self, parent):
        self.new_contact = None
        super().__init__(parent)

    def body(self, master):
        Label(master, text="Фамилия:").grid(row=0, sticky=W)
        Label(master, text="Имя:").grid(row=1, sticky=W)
        Label(master, text="Номер телефона:").grid(row=2, sticky=W)
        Label(master, text="Дата рождения:").grid(row=3, sticky=W)
        Label(master, text="Email:").grid(row=4, sticky=W)
        Label(master, text="Родственность:").grid(row=5, sticky=W)

        self.l_name_entry = Entry(master)
        self.f_name_entry = Entry(master)
        self.phone_entry = Entry(master)
        self.birthday_entry = Entry(master)
        self.email_entry = Entry(master)
        self.relationship_entry = Entry(master)

        self.l_name_entry.grid(row=0, column=1)
        self.f_name_entry.grid(row=1, column=1)
        self.phone_entry.grid(row=2, column=1)
        self.birthday_entry.grid(row=3, column=1)
        self.email_entry.grid(row=4, column=1)
        self.relationship_entry.grid(row=5, column=1)

    def apply(self):
        l_name = self.l_name_entry.get()
        f_name = self.f_name_entry.get()
        phone_number = self.phone_entry.get()
        birthday_date = self.birthday_entry.get()
        email_adress = self.email_entry.get()
        relationship_rodstvennost = self.relationship_entry.get()

        self.new_contact = Contact(None, l_name, f_name, phone_number, '', birthday_date, email_adress, '', relationship_rodstvennost)

class Main(Frame):
    alt = Contact(1, 'Zh', 'Alt', '87471800151', relationship='Коллега')
    alm = Contact(2, 'T', 'Alm', '87471800152', '', '15.09.1991', relationship='Друг')
    ali = Contact(3, 'Zh', 'Ali', '87471800153', '', '18.01.1995', 'ali@gmmail.com', 'well', relationship='Родственник')
    madi = Contact(4, 'Madi', 'Kaliyev', '87011910797','1910797@mail.ru',  relationship='студент')
    contacts = [alt, alm, ali, madi]

    def __init__(self, master):
        super().__init__(master)
        self.init_main()

    def init_main(self):
        self.toolbar = Frame(bg="#d7d8e0")
        self.toolbar.pack(side=TOP, fill=X)

        Button(self.toolbar, text="Добавить", command=self.add_contact).pack(side='left')
        Button(self.toolbar, text="Сохранить", command=self.save_to_cloud).pack(side='right')
        Button(self.toolbar, text="Изменить", command=self.edit_contact).pack(side='left')
        Button(self.toolbar, text="Удалить", command=self.delete_contact).pack(side='left')
        Button(self.toolbar, text="Открыть", command=self.load_contacts_from_excel).pack(side='right')
        Button(self.toolbar, text="Добавить фото", command=self.add_photo).pack(side='right')

        mainframe = Frame(bg="#e8e8e8")
        mainframe.pack(side=TOP, fill=BOTH)

        self.scrollbar = Scrollbar(mainframe)
        self.scrollbar.pack(side=RIGHT, fill=Y)

        self.listbox = Listbox(mainframe, listvariable=Variable(value=self.toValue()), width=50, height=20,
                                yscrollcommand=self.scrollbar.set)
        self.listbox.pack(side=LEFT, fill=BOTH)
        self.scrollbar.config(command=self.listbox.yview)

        self.right_frame = Frame(mainframe, bg="#e8e8e8")
        self.right_frame.pack(side=LEFT, fill=BOTH)

        self.listbox.bind('<<ListboxSelect>>', self.show_contact_details)

    def update_contact(self):
        selected_index = self.listbox.curselection()
        if selected_index:
            selected_contact = self.contacts[selected_index[0]]
            dialog = EditContactDialog(self, selected_contact)
            if dialog.result:
                self.refresh_listbox()

    def delete_contact(self):
        selected_index = self.listbox.curselection()
        if selected_index:
            confirm = messagebox.askyesno("Подтверждение", "Вы уверены, что хотите удалить выбранный контакт?")
            if confirm:
                del self.contacts[selected_index[0]]
                self.refresh_listbox()
                for widget in self.right_frame.winfo_children():
                    widget.destroy()
    def toValue(self) -> list:
        result_list = []
        for contact in self.contacts:
            result_list.append(contact.toStr())
        return result_list

    def add_contact(self):
        dialog = AddContactDialog(self)
        if dialog.new_contact:
            if not self.contact_exists(dialog.new_contact):
                self.contacts.append(dialog.new_contact)
                self.refresh_listbox()
            else:
                messagebox.showwarning("Повторяющийся контакт", "Контакт с такими данными уже существует.")

    def contact_exists(self, new_contact):
        for contact in self.contacts:
            if contact.l_name == new_contact.l_name and contact.f_name == new_contact.f_name \
                    and contact.phone_number == new_contact.phone_number:
                return True
        return False

    def load_contacts_from_excel(self):
        file_path = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Excel files", "*.xlsx;*.xls")])
        if file_path:
            try:
                wb = load_workbook(file_path)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    l_name, f_name, phone_number, birthday_date, email_adress, relationship_rodstvennost = row
                    new_contact = Contact(None, l_name, f_name, phone_number, '', birthday_date, email_adress, '', relationship_rodstvennost)
                    if not self.contact_exists(new_contact):
                        self.contacts.append(new_contact)
                self.refresh_listbox()
                messagebox.showinfo("Успех", "Контакты успешно загружены из файла Excel.")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Произошла ошибка при загрузке контактов: {e}")

    def save_to_cloud(self):
        try:
            connection = mysql.connector.connect(
                host='localhost',
                user='root',
                password='',
                database='python_db'
            )

            cursor = connection.cursor()

            cursor.execute('DELETE FROM contacts')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS contacts (
                id INT AUTO_INCREMENT PRIMARY KEY,
                l_name VARCHAR(255),
                f_name VARCHAR(255),
                phone_number VARCHAR(20),
                m_name VARCHAR(255),
                birthday VARCHAR(20),
                email VARCHAR(255),
                comment TEXT,
                relationship TEXT
            )
            ''')

            cursor.execute('''ALTER TABLE contacts MODIFY COLUMN relationship TEXT''')

            for contact in self.contacts:
                cursor.execute('''
                    INSERT INTO contacts (l_name, f_name, phone_number, m_name, birthday, email, comment, relationship)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ''', (contact.l_name, contact.f_name, contact.phone_number, contact.m_name,
                      contact.birthday, contact.email, contact.comment, contact.relationship))

            connection.commit()

            messagebox.showinfo("Успех", "Данные успешно сохранены в облако!")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка: {e}")
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()

    def refresh_listbox(self):
        self.listbox.delete(0, END)
        for contact in self.contacts:
            self.listbox.insert(END, contact.toStr())

    def add_photo(self):
        selected_index = self.listbox.curselection()
        if selected_index:
            contact = self.contacts[selected_index[0]]
            file_path = filedialog.askopenfilename()
            if file_path:
                with open(file_path, "rb") as image_file:
                    encoded_image = base64.b64encode(image_file.read())
                    contact.photo = encoded_image
                self.show_contact_details(None)

    def show_contact_details(self, event):
        selected_index = self.listbox.curselection()
        if selected_index:
            contact = self.contacts[selected_index[0]]
            self.display_contact_details(contact)

    def display_contact_details(self, contact):
        for widget in self.right_frame.winfo_children():
            widget.destroy()

        if contact.photo:
            img = ImageTk.PhotoImage(Image.open(io.BytesIO(base64.b64decode(contact.photo))))
            label = Label(self.right_frame, image=img)
            label.photo = img
            label.pack()
        else:
            label = Label(self.right_frame, text="No photo available")
            label.pack()

        details = f"Name: {contact.f_name} {contact.l_name}\nPhone: {contact.phone_number}\nEmail: {contact.email}\n" \
                  f"Birthday: {contact.birthday}\nComment: {contact.comment}\nRelationship: {contact.relationship}"
        details_label = Label(self.right_frame, text=details)
        details_label.pack()

    def edit_contact(self):
        selected_index = self.listbox.curselection()
        if selected_index:
            selected_contact = self.contacts[selected_index[0]]
            dialog = EditContactDialog(self, selected_contact)
            if dialog.result:
                self.refresh_listbox()

if __name__ == "__main__":
    window = Tk()
    Main(window).pack()
    window.title("Контакты")
    window.geometry("800x480")
    window.resizable(False, False)
    window.mainloop()