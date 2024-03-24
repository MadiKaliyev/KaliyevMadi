import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

def prodazhi(x):
    workbook = load_workbook(x)
    sheet = workbook.active
    data = [i[0] for i in sheet.iter_rows(values_only=True)]
    return data

data = prodazhi('889.xlsx')

chasi = np.arange(1, len(data) + 1)

plt.figure(figsize=(20, 10))
plt.bar(chasi, data, color='green')
plt.title('Количество продаж за каждый час')
plt.xlabel('Час')
plt.ylabel('Количество продаж')
plt.grid(axis='y')
plt.show()
